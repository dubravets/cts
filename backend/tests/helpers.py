from __future__ import annotations

import zipfile
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook


def make_excel_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlight"
    ws["A1"] = "Level"
    ws.merge_cells("A1:B1")
    ws["A2"] = "L1"
    ws["B2"] = 500
    ws["A3"] = "L2"
    ws["B3"] = 800
    raw = BytesIO()
    wb.save(raw)
    return raw.getvalue()


def make_word_bytes() -> bytes:
    document_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:p>
      <w:pPr><w:pStyle w:val="Heading1"/></w:pPr>
      <w:r><w:t>4 Display</w:t></w:r>
    </w:p>
    <w:p>
      <w:pPr><w:pStyle w:val="Heading2"/></w:pPr>
      <w:r><w:t>4.3 Backlight Mapping</w:t></w:r>
    </w:p>
    <w:tbl>
      <w:tr>
        <w:tc><w:p><w:r><w:t>Level</w:t></w:r></w:p></w:tc>
        <w:tc><w:p><w:r><w:t>Nits</w:t></w:r></w:p></w:tc>
      </w:tr>
      <w:tr>
        <w:tc><w:p><w:r><w:t>L1</w:t></w:r></w:p></w:tc>
        <w:tc><w:p><w:r><w:t>500</w:t></w:r></w:p></w:tc>
      </w:tr>
      <w:tr>
        <w:tc><w:p><w:r><w:t>L2</w:t></w:r></w:p></w:tc>
        <w:tc><w:p><w:r><w:t>800 nits</w:t></w:r></w:p></w:tc>
      </w:tr>
    </w:tbl>
    <w:sectPr/>
  </w:body>
</w:document>"""
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("word/document.xml", document_xml)
    return output.getvalue()


def upload_document(
    client: TestClient, *, doc_key: str, version: str, filename: str, payload: bytes
) -> str:
    response = client.post(
        "/api/v1/documents",
        data={"doc_key": doc_key, "version": version},
        files={"file": (filename, payload, "application/octet-stream")},
    )
    assert response.status_code == 200, response.text
    return response.json()["id"]


def create_reference(client: TestClient, *, doc_id: str, location: str, excerpt: str) -> str:
    response = client.post(
        "/api/v1/references",
        json={"doc_id": doc_id, "location": location, "excerpt": excerpt},
    )
    assert response.status_code == 200, response.text
    return response.json()["id"]


def create_spec_rule(
    client: TestClient,
    *,
    name: str,
    rule_type: str,
    expression: dict,
    reference_ids: list[str],
) -> str:
    response = client.post(
        "/api/v1/knowledge/spec-rules",
        json={
            "name": name,
            "rule_type": rule_type,
            "expression": expression,
            "reference_ids": reference_ids,
        },
    )
    assert response.status_code == 200, response.text
    return response.json()["id"]
