from io import BytesIO

from openpyxl import Workbook

from app.parsing.excel import parse_excel_bytes


def _build_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlight"

    ws["A1"] = "Level"
    ws.merge_cells("A1:B1")
    ws["A2"] = "L1"
    ws["B2"] = 500
    ws["A3"] = "Group-A"
    ws.merge_cells("A3:A4")
    ws["B4"] = "tail"

    raw = BytesIO()
    wb.save(raw)
    return raw.getvalue()


def test_parse_excel_expands_merged_cells_and_emits_references() -> None:
    rows = parse_excel_bytes(file_bytes=_build_workbook_bytes(), doc_id="doc-123")

    assert len(rows) == 4

    row1 = rows[0]
    assert row1.sheet_name == "Backlight"
    assert row1.row_index == 1
    assert row1.reference.doc_id == "doc-123"
    assert row1.reference.location == "sheet:Backlight!row:1"
    assert row1.reference.excerpt == "A1=Level | B1=Level"
    assert row1.cells[0].coordinate == "A1"
    assert row1.cells[0].source_coordinate == "A1"
    assert row1.cells[1].coordinate == "B1"
    assert row1.cells[1].source_coordinate == "A1"

    row4 = rows[3]
    assert row4.row_index == 4
    assert row4.reference.location == "sheet:Backlight!row:4"
    assert row4.cells[0].coordinate == "A4"
    assert row4.cells[0].source_coordinate == "A3"
    assert row4.cells[0].value == "Group-A"


def test_parse_excel_skips_empty_rows() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    ws["A3"] = "y"

    raw = BytesIO()
    wb.save(raw)
    rows = parse_excel_bytes(file_bytes=raw.getvalue(), doc_id="doc-1")

    assert [row.row_index for row in rows] == [1, 3]
