from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ParsedCell:
    coordinate: str
    source_coordinate: str
    value: str


@dataclass(frozen=True)
class ParsedReference:
    doc_id: str
    location: str
    excerpt: str


@dataclass(frozen=True)
class ParsedRow:
    sheet_name: str
    row_index: int
    cells: list[ParsedCell]
    reference: ParsedReference


def parse_excel_bytes(*, file_bytes: bytes, doc_id: str) -> list[ParsedRow]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    parsed_rows: list[ParsedRow] = []

    for sheet in workbook.worksheets:
        merged_source_map = _expand_merged_cells(sheet)
        parsed_rows.extend(
            _extract_sheet_rows(
                sheet=sheet,
                merged_source_map=merged_source_map,
                doc_id=doc_id,
            )
        )

    return parsed_rows


def parse_excel_file(*, file_path: str | Path, doc_id: str) -> list[ParsedRow]:
    raw = Path(file_path).read_bytes()
    return parse_excel_bytes(file_bytes=raw, doc_id=doc_id)


def _extract_sheet_rows(
    *, sheet: Worksheet, merged_source_map: dict[str, str], doc_id: str
) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    for row_index in range(1, sheet.max_row + 1):
        cells: list[ParsedCell] = []
        for col_index in range(1, sheet.max_column + 1):
            coord = f"{get_column_letter(col_index)}{row_index}"
            value = sheet.cell(row=row_index, column=col_index).value
            if value is None:
                continue

            normalized = _normalize_value(value)
            cells.append(
                ParsedCell(
                    coordinate=coord,
                    source_coordinate=merged_source_map.get(coord, coord),
                    value=normalized,
                )
            )

        if not cells:
            continue

        excerpt = " | ".join(f"{cell.coordinate}={cell.value}" for cell in cells)
        rows.append(
            ParsedRow(
                sheet_name=sheet.title,
                row_index=row_index,
                cells=cells,
                reference=ParsedReference(
                    doc_id=doc_id,
                    location=f"sheet:{sheet.title}!row:{row_index}",
                    excerpt=excerpt,
                ),
            )
        )
    return rows


def _expand_merged_cells(sheet: Worksheet) -> dict[str, str]:
    source_map: dict[str, str] = {}
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        source_coord = f"{get_column_letter(min_col)}{min_row}"
        source_value = sheet.cell(row=min_row, column=min_col).value

        sheet.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                sheet.cell(row=row, column=col, value=source_value)
                source_map[coord] = source_coord

    return source_map


def _normalize_value(value: object) -> str:
    if isinstance(value, str):
        return value.strip()
    return str(value)
